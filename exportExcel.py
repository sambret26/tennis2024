# -*- coding: utf-8 -*-

# Global packages
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from datetime import datetime
import openpyxl
import io

# My packages
from databases.repositories.playersRepository import PlayersRepository
from databases.repositories.matchsRepository import MatchsRepository
from databases.repositories.courtsRepository import CourtsRepository
from databases.repositories.teamsRepository import TeamsRepository
from databases.base import engine, session
from constants import constants


playersRepository = PlayersRepository(engine)
matchsRepository = MatchsRepository(engine)
courtsRepository = CourtsRepository(engine)
teamsRepository = TeamsRepository(engine)

def initiateMatchsSheet(sheet):
    for colNum, header in enumerate(constants.MATCHS_HEADERS, start=1): writeCell(sheet, colNum, 1, header)
    for colNum, width in enumerate(constants.MATCHS_COLUMN_WIDTHS, start=1): sheet.column_dimensions[openpyxl.utils.get_column_letter(colNum)].width = width
    matchsNames = getMatchsNamesInOrder()
    for rowNum, matchName in enumerate(matchsNames, start=3): addMatch(sheet, matchName, rowNum)
    addFormulas(sheet)

def initiateScheduleSheet(sheet):
    for colNum, header in enumerate(constants.SCHEDULE_HEADERS, start=1): writeCell(sheet, colNum, 1, header)
    for colNum, width in enumerate(constants.SCHEDULE_COLUMN_WIDTHS, start=1): sheet.column_dimensions[openpyxl.utils.get_column_letter(colNum)].width = width
    for day in range(15,31): addSchedule(sheet, f'{day}/06')

def initiatePlayersSheet(sheet):
    for colNum, header in enumerate(constants.PLAYERS_HEADERS, start=1): writeCell(sheet, colNum, 1, header)
    for colNum, width in enumerate(constants.PLAYERS_COLUMN_WIDTHS, start=1): sheet.column_dimensions[openpyxl.utils.get_column_letter(colNum)].width = width
    players = playersRepository.getPlayers(session)
    for rowNum, player in enumerate(players, start=2) :
        name = player.lastName + ' ' + player.firstName
        writeCell(sheet, 1, rowNum, name)
        writeCell(sheet, 2, rowNum, orZero(player.ranking))
    teams = teamsRepository.getTeams(session)
    for rowNum, team in enumerate(teams, start=2) :
        player1 = playersRepository.getPlayerNameById(session, team.player1Id)
        player2 = playersRepository.getPlayerNameById(session, team.player2Id)
        name = player1.lastName + constants.TEAM_SEPARATOR + player2.lastName
        writeCell(sheet, 4, rowNum, name)
        writeCell(sheet, 5, rowNum, orZero(team.ranking))
        writeCell(sheet, 6, rowNum, player1.firstName, None, Font(color="ffffff"))
        writeCell(sheet, 7, rowNum, player2.firstName, None, Font(color="ffffff"))

def centerAndBorderSheetAndWhiteFont(sheet, maxColumn = None):
    borderType = Side(style='thin')
    border = Border(top=borderType, bottom=borderType, left=borderType, right=borderType)
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column if maxColumn == None else maxColumn):
        for cell in row:
            cell.alignment = Alignment(vertical="center", horizontal="center")
            if cell.value != None or cell.coordinate in sheet.merged_cells : cell.border = border
    addWhiteFont(sheet)

def centerAndBorderSheetAndWhiteFontMatchs(sheet):
    borderType = Side(style='thin')
    border = Border(top=borderType, bottom=borderType, left=borderType, right=borderType)
    row = 1
    while True and row < 200:
        if row == 2  :
            row = 3
            continue
        if sheet.cell(column = 1, row = row).value == None : break
        for col in range(1, 10):
            sheet.cell(column = col, row = row).alignment = Alignment(vertical="center", horizontal="center")
            sheet.cell(column = col, row = row).border = border
        row += 1
    addWhiteFont(sheet)

def addMatch(sheet, matchName, rowNum):
    player1Id, player2Id, winnerId, score, panel, nextRound = matchsRepository.getMatchInfosByName(session, matchName)
    player1 = getPlayer(matchName, rowNum, player1Id)
    player2 = getPlayer(matchName, rowNum, player2Id)
    writeCell(sheet, 1, rowNum, matchName)
    writeCell(sheet, 2, rowNum, player1)
    writeCell(sheet, 4, rowNum, player2)
    writeCell(sheet, 6, rowNum, orZero(panel))
    if(player1Id and winnerId == player1Id): writeCell(sheet, 7, rowNum, player1)
    elif(player2Id and winnerId == player2Id): writeCell(sheet, 7, rowNum, player1)
    else: writeCell(sheet, 7, rowNum, '')
    writeCell(sheet, 8, rowNum, orZero(score))
    writeCell(sheet, 9, rowNum, orZero(nextRound))

def getPlayer(matchName, rowNum, playerId):
    if not playerId : return None
    if playerId.startswith("VS") or playerId.startswith("VD"):
        case = getCase(playerId, matchName, rowNum)
        if case : return constants.PLAYER_NAME_FORMULA.replace("CASE", case).replace("PLAYER", playerId)
        return playerId
    if playerId.startswith("VT") or playerId.startswith("VP") or playerId.startswith("QE"):
        return playerId
    if matchName.startswith("S"):
        player =  playersRepository.getPlayerNameById(session, playerId)
        if player : return f"{player.lastName} {player.firstName}"
        return None
    team = teamsRepository.getTeamNameById(session, playerId)
    player1 = playersRepository.getPlayerNameById(session, team.player1Id)
    player2 = playersRepository.getPlayerNameById(session, team.player2Id)
    if player1 and player2 : return f"{player1.lastName}--{player2.lastName}"
    return None

def getCase(playerId, matchName, rowNum):
    cat1 = playerId[1:3]
    cat2 = matchName[0:2]
    if cat2 != cat2: return None
    number1 = int(playerId[-2:])
    number2 = int(matchName[-2:])
    return "G"+str(rowNum - number2 + number1)

def addFormulas(sheet):
    for rowNum in range (3,200):
        writeCell(sheet, 3, rowNum, constants.MATCHS_FORMULA_COLUMN_C.replace("ROWNUM", str(rowNum)))
        writeCell(sheet, 5, rowNum, constants.MATCHS_FORMULA_COLUMN_E.replace("ROWNUM", str(rowNum)))

def orZero(value):
    return value if value !=None else 0

def addSchedule(sheet, date):
    court1 = courtsRepository.getCourtByName(session, "1")
    court2 = courtsRepository.getCourtByName(session, "2")
    court3 = courtsRepository.getCourtByName(session, "3")
    matchsCourt1 = matchsRepository.getMatchsByDateAndCourt(session, date, court1.id)
    matchsCourt2 = matchsRepository.getMatchsByDateAndCourt(session, date, court2.id)
    matchsCourt3 = matchsRepository.getMatchsByDateAndCourt(session, date, court3.id)
    sortedMatchsCourt1 = sorted(matchsCourt1, key=lambda match: inMinutes(match.hour))
    sortedMatchsCourt2 = sorted(matchsCourt2, key=lambda match: inMinutes(match.hour))
    sortedMatchsCourt3 = sorted(matchsCourt3, key=lambda match: inMinutes(match.hour))
    if isNotPast(date):
        timeSlots1 = addSlot(sortedMatchsCourt1, date, court1.id)
        timeSlots2 = addSlot(sortedMatchsCourt2, date, court2.id)
    timeSlots = timeSlots1 + timeSlots2 + sortedMatchsCourt3
    if len(timeSlots) == 0: return
    timeSlots.sort(key=lambda ts: inMinutes(ts[0]))
    firstRow = sheet.max_row + 1
    previousHour = None
    toMerge = False
    for rowNum, timeSlot in enumerate(timeSlots, start=firstRow):
        hour, match, courtId = timeSlot
        courtName = courtsRepository.getCourtNameById(session, courtId)
        rowNum = sheet.max_row + 1
        writeCell(sheet, 1, rowNum, changeDate(date))
        writeCell(sheet, 2, rowNum, hour)
        writeCell(sheet, 3, rowNum, courtName)
        writeCell(sheet, 4, rowNum, match)
        for column in range(5,13): writeCell(sheet, column, rowNum, constants.SCHEDULE_FORMULA.replace("ROWNUM", str(rowNum)).replace("COLUMN", str(column-3)))
        if hour == previousHour and not toMerge:
            toMerge = True
            start_row = rowNum -1
        elif hour != previousHour and toMerge:
            sheet.merge_cells(start_row=start_row, end_row=rowNum -1, start_column=2, end_column=2)
            toMerge=False
        previousHour = hour
    if toMerge: sheet.merge_cells(start_row=start_row, end_row=rowNum, start_column=2, end_column=2)
    sheet.merge_cells(start_row=firstRow, end_row=firstRow + len(timeSlots) - 1, start_column=1, end_column=1)

def isNotPast(date):
    return int(datetime.now().strftime("%m%d")) < 100 * int(date[3:]) + int(date[0:2])

def addSlot(timeSlot, date, court):
    hours = [hour for hour, name, court in timeSlot]
    href = getHref(date)
    hours.insert(0, href)
    max = 10
    for index, hour in enumerate(hours):
        if max == 0 : return timeSlot
        max = max -1
        if len(hours) > index+1 : nextHour = hours[index+1]
        else : nextHour = constants.MAX_HOUR
        while findDiff(hour, nextHour) > 179 :
            nextHour = plus90(hour)
            if inMinutes(nextHour) > inMinutes(href) + 89 :
                hours.insert(index +1, nextHour)
                timeSlot.insert(index, (nextHour, 0, court))
            else :
                hours.insert(index +1, href)
                break
    return timeSlot

def getHref(date):
    day = int(date.split("/")[0])
    if day in [15,16,22,23,29,30] : return constants.WEEKEND_START_HOUR
    return constants.WEEK_START_HOUR

def findDiff(hour, nextHour):
    return inMinutes(nextHour) - inMinutes(hour)

def inMinutes(hour):
    split = hour.upper().split("H")
    if len(split) == 1 :
        h = split[0]
    else :
        h,m = split
    h = int(h)
    if m == '': m = 0
    else: m = int(m)
    return h*60+m

def plus90(hour):
    minutes = inMinutes(hour)
    minutes = minutes+90
    h = minutes//60
    m = (minutes - 60*h)
    if m ==0 : return f"{h}H"
    return f"{h}H{m}"

def changeDate(date):
    try:
        date += '/2024'
        parsed_date = datetime.strptime(date, '%d/%m/%Y')
        formatted_date = parsed_date.strftime('%A %d %B')
        return formatted_date.title()
    except ValueError:
        return date

def addWhiteFont(sheet):
    rule = CellIsRule(operator='equal', formula=['0'], font=Font(color="ffffff"))
    sheet.conditional_formatting.add('A1:ZZ1048576', rule)

def addCourtColor(sheet):
    rule1 = FormulaRule(formula=['AND(OR($C2=2,$C2="2"),D2<>0)'], fill=PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"))
    rule2 = FormulaRule(formula=['AND(OR($C2=2,$C2="2"),D2=0)'], fill=PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"), font=Font(color="D9D9D9"))
    rule3 = FormulaRule(formula=['AND(OR($C2=3,$C2="3"),D2<>0)'], fill=PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid"))
    rule4 = FormulaRule(formula=['AND(OR($C2=3,$C2="3"),D2=0)'], fill=PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid"), font=Font(color="A6A6A6"))
    for rule in [rule1, rule2, rule3, rule4]:
        sheet.conditional_formatting.add('D2:L1048576', rule)

def getMatchsNamesInOrder():
    allMatchs = []
    for categorie in constants.CATEGORIES:
        matchs = matchsRepository.getMatchsNamesByCategories(session, categorie)
        for match in matchs: allMatchs.append(match.name)
    return allMatchs

def initiateExcel():
    file = openpyxl.Workbook()
    defaultSheet = file['Sheet']
    file.remove(defaultSheet)
    matchSheet = file.create_sheet(constants.MATCH_SHEET_NAME)
    scheduleSheet = file.create_sheet(constants.SCHEDULE_SHEET_NAME)
    playersSheet = file.create_sheet(constants.PLAYERS_SHEET_NAME)
    initiatePlayersSheet(playersSheet)
    initiateMatchsSheet(matchSheet)
    initiateScheduleSheet(scheduleSheet)
    addCourtColor(scheduleSheet)
    centerAndBorderSheetAndWhiteFontMatchs(matchSheet)
    centerAndBorderSheetAndWhiteFont(scheduleSheet)
    centerAndBorderSheetAndWhiteFont(playersSheet, 5)
    excelBytes = io.BytesIO()
    file.save(excelBytes)
    excelBytes.seek(0)
    return excelBytes

def writeCell(sheet, column, row, value, fill=None, font=None):
    if value in ['0', 0, '', None] : value = ''
    sheet.cell(column=column, row=row, value=value)
    if fill : sheet.cell(column=column, row=row).fill = fill
    if font : sheet.cell(column=column, row=row).font = font
