# -*- coding: utf-8 -*-

# Global packages
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from datetime import datetime
import openpyxl
import discord
import locale
import io

#My packages
from databases.repositories.playersRepository import PlayersRepository
from databases.repositories.matchsRepository import MatchsRepository
from databases.repositories.courtsRepository import CourtsRepository
from databases.repositories.teamsRepository import TeamsRepository
from databases.base import engine, session
from constants import constants
import update
import moja

playersRepository = PlayersRepository(engine)
matchsRepository = MatchsRepository(engine)
courtsRepository = CourtsRepository(engine)
teamsRepository = TeamsRepository(engine)

async def readExcel(message):
    data = await message.attachments[0].read()
    fileData = io.BytesIO(data)
    inputFile = openpyxl.load_workbook(fileData)
    if not await checkExcel(inputFile): return
    listMatchsToAdd = []
    listMatchsToUpdate = []
    listMatchsNamesToRemove = []
    listErrors = []
    handleMatchsSheet(inputFile[constants.MATCH_SHEET_NAME], listMatchsToAdd, listMatchsToUpdate, listMatchsNamesToRemove)
    handleScheduleSheet(inputFile[constants.SCHEDULE_SHEET_NAME], listMatchsToAdd, listMatchsToUpdate, listMatchsNamesToRemove, listErrors)
    reportFile = createReportFile()
    fillReportFileAndUpdateDB(reportFile, listMatchsToAdd, listMatchsToUpdate, listMatchsNamesToRemove, listErrors)
    if listErrors != [] : addErrorsReportSheet(reportFile, listErrors)
    formattingReportFile(reportFile)
    excelBytes = io.BytesIO()
    reportFile.save(excelBytes)
    excelBytes.seek(0)
    await message.channel.send(file=discord.File(fp=excelBytes, filename = constants.REPORT_FILENAME))

async def checkExcel(inputFile):
    for sheetName in [constants.PLAYERS_SHEET_NAME, constants.MATCH_SHEET_NAME, constants.SCHEDULE_SHEET_NAME]:
        if not sheetName in inputFile.sheetnames :
            await message.channel.send(constants.SHEET_NOT_FOUND.replae("SHEET", constants.sheetName))
            return False
    return True

def handleMatchsSheet(sheet, listToAdd, listToUpdate, listToRemove):
    row = 3
    listHandled = []
    while True:
        matchName = readCell(sheet, 1, row)
        if matchName is None: break
        player1 = readCell(sheet, 2, row)
        player2 = readCell(sheet, 4, row)
        panel = readCell(sheet, 6, row)
        winner = readCell(sheet, 7, row)
        score = readCell(sheet, 8, row)
        nextRound = readCell(sheet, 9, row)
        player1Id = getPlayerIdByNameAndMatchType(player1, matchName)
        player2Id = getPlayerIdByNameAndMatchType(player2, matchName)
        winnerId = getWinnerId(winner, player1, player1Id, player2, player2Id)
        newMatch = [player1Id, player2Id, None, None, None, winnerId, score if score != 0 else None, panel if panel != 0 else None, nextRound if nextRound != 0 else None]
        row += 1
        matchInfosInDB = matchsRepository.getMatchInfosByName(session, matchName)
        if matchInfosInDB is None:
            listToAdd.append([matchName, player1Id, player1, player2Id, player2, None, None, None, winner, winnerId, score, panel, nextRound])
            continue
        listHandled.append(f'{matchName}')
        player1IdInDB = matchInfosInDB[0]
        player2IdInDB = matchInfosInDB[1]
        winnerIdInDB = matchInfosInDB[2]
        scoreInDB = matchInfosInDB[3]
        panelInDB = matchInfosInDB[4]
        nextRoundInDB = matchInfosInDB[5]
        oldMatch = [player1IdInDB, player2IdInDB, None, None, None, winnerIdInDB, scoreInDB, panelInDB, nextRoundInDB]
        if oldMatch != newMatch: listToUpdate.append([matchName, oldMatch, newMatch])
    listNamesToRemove = matchsRepository.getMatchsNamesToDelete(session, listHandled)
    for m in listNamesToRemove: listToRemove.append(m)

def handleScheduleSheet(sheet, listToAdd, listToUpdate, listNamesToRemove, listErrors):
    row = 2
    listHandled = []
    while True:
        date = readCellMaybeMerge(sheet, 1, row)
        if date is None: break
        hour = readCellMaybeMerge(sheet, 2, row)
        if hour is None: break
        matchName = readCell(sheet, 4, row)
        if matchName in ['0', 0, ' ', '', None]:
            row+=1
            continue
        listHandled.append(f'{matchName}')
        if matchName in [row[0] for row in listNamesToRemove]:
            listErrors.append(f"Erreur : Le match {matchName} est programmé mais doit être supprimé")
            row+=1
            continue
        courtName = readCell(sheet, 3, row)
        courtId = courtsRepository.getCourtIdByName(session, courtName)
        row+=1
        newProgram = [reverseChangeDate(date), hour, courtId]
        programInDB = matchsRepository.getMatchProgramByName(session, matchName)
        if programInDB is None:
            if matchName in [row[0] for row in listToAdd]: addPgInListAdd(matchName, newProgram, listToAdd)
            else : listErrors.append(f"Erreur : Le match {matchName} est programmé mais n'existe pas en base de données")
            continue
        oldProgram = [programInDB[0], programInDB[1], programInDB[2]]
        if matchName in [row[0] for row in listToUpdate]:
            addOldAndNewPgInListUpdate(matchName, oldProgram, newProgram, listToUpdate)
            continue
        if oldProgram != newProgram: addRescheduledMatchInListUpdate(matchName, newProgram, listToUpdate)
    unscheduleMatchs(matchsRepository.getMatchsToUnschedule(session, listHandled), listToUpdate, listNamesToRemove)

def unscheduleMatchs(matchsNameList, listToUpdate, listNamesToRemove):
    for matchName in matchsNameList:
        match = matchsRepository.getAllMatchInfosByName(session, matchName)
        if matchName in listNamesToRemove: continue
        if (match.day, match.hour, match.courtId) == (None, None, None) : continue
        courtId = courtsRepository.getCourtIdByName(session, match.courtId)
        if matchName in [row[0] for row in listToUpdate]:
            oldProgram = [match.day, match.hour, courtId]
            newProgram = [None, None, None]
            addOldAndNewPgInListUpdate(matchName, oldProgram, newProgram, listToUpdate)
            continue
        oldMatch = [match.player1Id, match.player2Id, match.day, match.hour, courtId, match.winnerId, match.score, match.panel, match.nextRound]
        newMatch = [match.player1Id, match.player2Id, None, None, None, match.winnerId, match.score, match.panel, match.nextRound]
        listToUpdate.append([matchName, oldMatch, newMatch])

def addPgInListAdd(matchName, program, listToAdd):
    for index, matchInList in enumerate(listToAdd):
        if matchInList[0] == matchName:
            listToAdd[index][5] = program[0]
            listToAdd[index][6] = program[1]
            listToAdd[index][7] = program[2]
            return

def addOldAndNewPgInListUpdate(matchName, oldProgram, newProgram, listToUpdate):
    for index , matchInList in enumerate(listToUpdate):
        if matchInList[0] == matchName:
            listToUpdate[index][1][2] = oldProgram[0]
            listToUpdate[index][1][3] = oldProgram[1]
            listToUpdate[index][1][4] = oldProgram[2]
            listToUpdate[index][2][2] = newProgram[0]
            listToUpdate[index][2][3] = newProgram[1]
            listToUpdate[index][2][4] = newProgram[2]
            return

def addRescheduledMatchInListUpdate(matchName, newProgram, listToUpdate):
    match = matchsRepository.getAllMatchInfosByName(session, matchName)
    courtId = courtsRepository.getCourtIdByName(session, match.courtId)
    oldMatch = [match.player1Id, match.player2Id, match.day, match.hour, courtId, match.winnerId, match.score, match.panel, match.nextRound]
    newMatch = [match.player1Id, match.player2Id, newProgram[0], newProgram[1], newProgram[2], match.winnerId, match.score, match.panel, match.nextRound]
    listToUpdate.append([matchName, oldMatch, newMatch])

def readCellMaybeMerge(sheet, column, row):
    if sheet.cell(column=column, row=row).coordinate in sheet.merged_cells :
        while row > 0:
            value = readCell(sheet, column, row)
            if value != None : return value
            row -= 1
        return None
    return readCell(sheet, column, row)

def readCell(sheet, column, row):
    return sheet.cell(column=column, row=row).value

def orZero(value):
    return value if value !=None else 0

def changeDate(date):
    if date is None: return None
    try:
        locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8')
        newDate = date + '/2024'
        parsed_date = datetime.strptime(newDate, '%d/%m/%Y')
        formatted_date = parsed_date.strftime('%A %d %B')
        return formatted_date.title()
    except ValueError:
        return date

def reverseChangeDate(formatted_date):
    try:
        locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8')
        parsed_date = datetime.strptime(formatted_date, '%A %d %B')
        reversed_date = parsed_date.strftime('%d/%m')
        return reversed_date
    except ValueError:
        return formatted_date

def getPlayerIdByNameAndMatchType(player, matchName):
    if player in [" ", "", None] : return None
    if player.startswith("VS") or player.startswith("VD") or player.startswith("VT") or player.startswith("VP") or player.startswith("QE"):
        return player
    if player.startswith("=IF"):
        return player.split('\"')[1]
    if matchName.startswith("S"):
        split = player.split(' ')
        playerName = ' '.join(split[0:-1])
        playerFirstName = split[-1]
        return str(playersRepository.getPlayerIdByName(session, playerName, playerFirstName))
    return str(getTeamIdByName(player))

def getPlayerNameByIdAndMatchType(playerId, matchName):
    if not playerId or playerId == "None" : return None
    if playerId.startswith("V"): return playerId
    if matchName.startswith("S"):
        player = playersRepository.getPlayerNameById(session, playerId)
        if player : return f"{player.lastName} {player.firstName}"
        return None
    team = teamsRepository.getTeamNameById(session, playerId)
    player1 = playersRepository.getPlayerNameById(session, team.player1Id)
    player2 = playersRepository.getPlayerNameById(session, team.player2Id)
    if player1 and player2 : return f"{player1.lastName}--{player2.lastName}"
    return None

def getTeamIdByName(name):
    teams = teamsRepository.getTeams(session)
    for team in teams:
        player1 = playersRepository.getPlayerNameById(session, team.player1Id)
        player2 = playersRepository.getPlayerNameById(session, team.player2Id)
        if [player1.lastName, player2.lastName] == (name.split(constants.TEAM_SEPARATOR)):
            return team.id
    return None

def getWinnerId(winner, player1, player1Id, player2, player2Id):
    if winner == player1: return player1Id
    if winner == player2: return player2Id
    return None

def createReportFile():
    file = openpyxl.Workbook()
    defaultSheet = file['Sheet']
    file.remove(defaultSheet)
    matchsReportSheet = file.create_sheet(constants.MATCHS_REPORT_SHEET)
    initiateMatchsReportSheet(matchsReportSheet)
    return file

def initiateMatchsReportSheet(sheet):
    for colNum, width in enumerate([7, 27, 27, 20, 12, 10, 27, 18, 8, 10], start=1): sheet.column_dimensions[openpyxl.utils.get_column_letter(colNum)].width = width
    writeCell(sheet, 1, 1, "Matchs ajoutés")
    writeCell(sheet, 1, 5, "Matchs modifiés")
    writeCell(sheet, 1, 9, "Matchs supprimés")
    for r in [2,6,10] :
        writeCell(sheet, 1, r, "Nom")
        writeCell(sheet, 2, r, "Joueur 1")
        writeCell(sheet, 3, r, "Joueur 2")
        writeCell(sheet, 4, r, "Heure")
        writeCell(sheet, 5, r, "Nom")
        writeCell(sheet, 6, r, "Court")
        writeCell(sheet, 7, r, "Vainqueur")
        writeCell(sheet, 8, r, "Score")
        writeCell(sheet, 9, r, "Tableau")
        writeCell(sheet, 10, r, "Suivant")

def fillReportFileAndUpdateDB(reportFile, listMatchsToAdd, listMatchsToUpdate, listMatchsNamesToRemove, listErrors):
    for match in listMatchsToAdd: addMatch(reportFile, match)
    for match in listMatchsToUpdate: updateMatch(reportFile, match, listErrors)
    for match in listMatchsNamesToRemove: removeMatch(reportFile, match, listErrors)

def addMatch(reportFile, match):
    sheet = reportFile[constants.MATCHS_REPORT_SHEET]
    row = 2
    while True:
        if readCell(sheet, 1, row) is None: break
        row+=1
    green = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type='solid')
    white = Font(color = "ffffff")
    sheet.insert_rows(row)
    writeCell(sheet, 1, row, match[0], green, white) #Name
    writeCell(sheet, 2, row, match[2], green, white) #Player1
    writeCell(sheet, 3, row, match[4], green, white) #Player2
    writeCell(sheet, 4, row, changeDate(match[5]), green, white) #Date
    writeCell(sheet, 5, row, match[6], green, white) #Hour
    writeCell(sheet, 6, row, match[7], green, white) #Court
    writeCell(sheet, 7, row, match[8], green, white) #Winner
    writeCell(sheet, 8, row, match[10], green, white) #Score
    writeCell(sheet, 9, row, match[11], green, white) #Panel
    writeCell(sheet, 10, row, match[12], green, white) #NextRound
    matchsRepository.insertMatch(session, match[0][0:2], match[0], match[1], match[3], match[5], match[6], match[7], match[9], match[10], match[11], match[12])

def updateMatch(reportFile, match, listErrors):
    matchName = match[0]
    oldMatch = match[1]
    newMatch = match[2]
    matchId = matchsRepository.getMatchIdByName(session, matchName)
    sheet = reportFile[constants.MATCHS_REPORT_SHEET]
    row = 4
    while True and row < 100:
        if readCellMaybeMerge(sheet, 1, row) == "Matchs modifiés": break
        row += 1
    if row == 100 :
        listErrors.append("Ligne 'Matchs modifiés' non trouvées")
        return
    while True:
        if readCellMaybeMerge(sheet, 1, row) is None: break
        row += 1
    orange = PatternFill(start_color='FFFFA500', end_color='FFFFA500', fill_type='solid')
    white = Font(color = "ffffff")
    sheet.insert_rows(row)
    sheet.insert_rows(row)
    writeCell(sheet, 1, row, matchName)
    sheet.merge_cells(start_row=row, end_row=row+1, start_column=1, end_column=1)
    newWinner = None
    oldWinner = None
    toSchedule = False
    for index in range(0, 2):
        oldPlayerName = getPlayerNameByIdAndMatchType(oldMatch[index], matchName)
        if oldMatch[5] == oldMatch[index] : oldWinner = oldPlayerName
        if oldMatch[index] != newMatch[index]:
            newPlayerName = getPlayerNameByIdAndMatchType(newMatch[index], matchName)
            if newMatch[5] == newMatch[index] : newWinner = newPlayerName
            if index == 0: setPlayer1(matchId, newMatch[index], oldMatch[index])
            else: setPlayer2(matchId, newMatch[index], oldMatch[index])
            writeCell(sheet, 2+index, row, oldPlayerName, orange, white)
            writeCell(sheet, 2+index, row+1, newPlayerName, orange, white)
        else:
            writeCell(sheet, 2+index, row, oldPlayerName)
            sheet.merge_cells(start_row=row, end_row=row+1, start_column=2+index, end_column=2+index)
    for index in range (2, 9):
        if oldMatch[index] != newMatch[index]:
            if index == 2:
                matchsRepository.setDay(session, matchId, newMatch[index])
                newMatch[index] = changeDate(newMatch[index])
                oldMatch[index] = changeDate(oldMatch[index])
                toSchedule = True
            elif index == 3:
                matchsRepository.setHour(session, matchId, newMatch[index])
                toSchedule = True
            elif index == 4: setCourt(matchId, newMatch[index])
            elif index == 5: setWinner(matchId, newMatch[index])
            elif index == 6: setScore(matchId, newMatch[index])
            elif index == 7: matchsRepository.setPanel(session, matchId, newMatch[index]) #TODO ?
            else : matchsRepository.setNextRound(session, matchId, newMatch[index]) #TODO ?
            if index == 5:
                writeCell(sheet, 2+index, row, oldWinner, orange, white)
                writeCell(sheet, 2+index, row+1, newWinner, orange, white)
            else:
                writeCell(sheet, 2+index, row, oldMatch[index], orange, white)
                writeCell(sheet, 2+index, row+1, newMatch[index], orange, white)
        else:
            if index == 5 : writeCell(sheet, 2+index, row, oldWinner)
            else : writeCell(sheet, 2+index, row, oldMatch[index])
            sheet.merge_cells(start_row=row, end_row=row+1, start_column=2+index, end_column=2+index)
    if toSchedule : update.updateEvent(matchId)

def setPlayer1(matchId, playerId, oldPlayerId):
    matchFftId = matchsRepository.getFftIdById(session, matchId)
    player = playersRepository.getFftIdAndInscriptionIdById(session, playerId)
    if player :
        if moja.setPlayer(matchFftId, player, "A"):
            matchsRepository.setPlayer1(session, matchId, playerId)
    else :
        player = playersRepository.getFftIdAndInscriptionIdById(session, oldPlayerId)
        if moja.deletePlayer(matchFftId, player, "A"):
            matchsRepository.setPlayer1(session, matchId, playerId)

def setPlayer2(matchId, playerId, oldPlayerId):
    matchFftId = matchsRepository.getFftIdById(session, matchId)
    player = playersRepository.getFftIdAndInscriptionIdById(session, playerId)
    if player :
        if moja.setPlayer(matchFftId, player, "B"):
            matchsRepository.setPlayer2(session, matchId, playerId)
    else :
        player = playersRepository.getFftIdAndInscriptionIdById(session, oldPlayerId)
        if moja.deletePlayer(matchFftId, player, "B"):
            matchsRepository.setPlayer2(session, matchId, playerId)

def setCourt(matchId, courtName):
    courtId = courtsRepository.getCourtIdByName(session, courtName)
    matchsRepository.setCourt(session, matchId, courtId)
    #matchFftId = matchsRepository.getFftIdById(session, matchId)
    #moja.setCourt(matchFftId, court)

def setWinner(matchId, winnerId):
    matchsRepository.setWinner(session, matchId, winnerId)
    #matchFftId = matchsRepository.getFftIdById(session, matchId)
    #moja.setWinner(matchFftId, winnerId)

def setScore(matchId, score):
    matchsRepository.setScore(session, matchId, score)
    #matchFftId = matchsRepository.getFftIdById(session, matchId)
    #moja.setScore(matchFftId, score)

def removeMatch(reportFile, matchName, listErrors):
    match = matchsRepository.getAllMatchInfosByName(session, matchName)
    player1 = getPlayerNameByIdAndMatchType(match.player1Id, matchName)
    player2 = getPlayerNameByIdAndMatchType(match.player2Id, matchName)
    courtName = courtsRepository.getCourtNameById(sessio, match.courtId)
    if match.winnerId == match.player1Id : winner = player1
    elif match.winnerId== match.player2Id : winner = player2
    else: winner = None
    sheet = reportFile[constants.MATCHS_REPORT_SHEET]
    row = 4
    while True and row < 100:
        if readCellMaybeMerge(sheet, 1, row) == "Matchs supprimés" : break
        row += 1
    if row == 100 :
        listErrors.append("Ligne 'Match supprimés' non trouvée")
        return
    while True:
        if readCellMaybeMerge(sheet, 1, row) is None : break
        row += 1
    red = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type='solid')
    white = Font(color = "ffffff")
    writeCell(sheet, 1, row, matchName, red, white) #Name
    writeCell(sheet, 2, row, player1, red, white) #Player1
    writeCell(sheet, 3, row, player2, red, white) #Player2
    writeCell(sheet, 4, row, orZero(changeDate(match.day)), red, white) #Date
    writeCell(sheet, 5, row, orZero(match.hour), red, white) #Hour
    writeCell(sheet, 6, row, orZero(courtName), red, white) #Court
    writeCell(sheet, 7, row, winner, red, white) #Winner
    writeCell(sheet, 8, row, orZero(match.score), red, white) #Score
    writeCell(sheet, 9, row, orZero(match.panel), red, white) #Panel
    writeCell(sheet, 10, row, orZero(match.nextRound), red, white) #NextRound
    matchsRepository.deleteMatchById(session, match.id)

def addErrorsReportSheet(reportFile, listErrors):
    errorsReportSheet = reportFile.create_sheet(constants.ERRORS_REPORT_SHEET)
    errorsReportSheet.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 50
    writeCell(errorsReportSheet, 1, 1, "Liste des erreurs rencontrées")
    row = 2
    for error in listErrors:
        writeCell(errorsReportSheet, 1, row, error)
        row += 1

def formattingReportFile(reportFile):
    for sheet in reportFile:
        borderType = Side(style='thin')
        border = Border(top=borderType, bottom=borderType, left=borderType, right=borderType)
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.alignment = Alignment(vertical="center", horizontal="center")
                if cell.value != None or cell.coordinate in sheet.merged_cells : cell.border = border
    sheet = reportFile[constants.MATCHS_REPORT_SHEET]
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value in ["Matchs ajoutés", "Matchs modifiés", "Matchs supprimés"] : sheet.merge_cells(start_row=cell.row, end_row=cell.row, start_column=1, end_column=10)

def writeCell(sheet, column, row, value, fill=None, font=None):
    if value in ['0', 0, '', None] : value = ' '
    sheet.cell(column=column, row=row, value=value)
    if fill : sheet.cell(column=column, row=row).fill = fill
    if font : sheet.cell(column=column, row=row).font = font
